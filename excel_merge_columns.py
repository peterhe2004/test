import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('pdtest1.xlsx')

# Select the desired sheet
sheet = workbook['pdtest1']  # Replace 'Sheet1' with your sheet name

# Define the columns you want to merge
column1 = 'B'  # Replace 'A' with the desired column letter
column2 = 'C'  # Replace 'B' with the desired column letter

# Get the maximum row count in the columns
# max_row = max(sheet.max_row, sheet[column1].max_row, sheet[column2].max_row)

# Iterate over the rows and merge the values
for row in range(1, 6):
    cell1 = sheet[column1 + str(row)]
    cell2 = sheet[column2 + str(row)]
    merged_value = f"{cell1.value}{cell2.value}"
    sheet[column1 + str(row)].value = merged_value
    sheet[column2 + str(row)].value = None

# Save the modified workbook
workbook.save('pdtest2.xlsx')
